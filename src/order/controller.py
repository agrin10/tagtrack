from src import db
from src.order.models import Order, OrderImage, OrderValue , OrderFile , FormNumberSequence , Customer
from src.production.models import JobMetric, Machine, ProductionStepLog
from flask_login import current_user
from datetime import datetime, date
from typing import Tuple, Dict, Any, List , Optional
import traceback
from sqlalchemy import func
import io
import os
import uuid
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import re
import uuid , logging
from itertools import zip_longest
from src.utils import parse_date_input

logging.basicConfig(level=logging.INFO)
# Add these constants at the top of the file
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB
UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'static', 'uploads', 'orders')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

def _get_next_form_number_for_year() -> int:
    current_year = datetime.now().year
    start_of_year = datetime(current_year, 1, 1)
    end_of_year = datetime(current_year, 12, 31, 23, 59, 59)

    # Query for the maximum form_number within the current year
    max_form_number = db.session.query(func.max(Order.form_number)).filter(
        Order.created_at.between(start_of_year, end_of_year)
    ).scalar()

    return (max_form_number or 0) + 1

from sqlalchemy import select
from sqlalchemy.exc import SQLAlchemyError

def allocate_form_number_for_year(start_from: int | None = None) -> int:
    """
    Atomically allocate the next form number for the current year.
    If start_from is provided and greater than the stored last_number,
    set and return start_from (so the first allocated number can start at this).
    This function assumes it's called within an active db.session transaction.
    """
    current_year = datetime.now().year

    # Lock the row for this year (for update)
    seq = db.session.query(FormNumberSequence).filter_by(year=current_year).with_for_update().first()

    if not seq:
        # create a new sequence row
        initial_last = 0
        seq = FormNumberSequence(year=current_year, last_number=initial_last)
        db.session.add(seq)
        # flush so DB assigns id and the row exists for locking
        db.session.flush()

    # Now decide next number
    if start_from and start_from > seq.last_number:
        next_number = start_from
        seq.last_number = start_from  # set the last_number to start_from (first allocated)
    else:
        seq.last_number = seq.last_number + 1
        next_number = seq.last_number

    # flush to persist updated last_number (but do NOT commit here; commit will happen in caller)
    db.session.flush()
    return next_number


def add_order(form_data: Dict[str, Any], files=None) -> Tuple[bool, Dict[str, Any]]:
    try:
        # Debug prints (optional)
        print("Raw form data received:", form_data)
        print("Files received:", files)

        # Normalize form_data values (convert numeric like int/float to str and strip whitespace)
        form_data = {
            k: (str(v).strip() if isinstance(v, (str, int, float)) else v)
            for k, v in (form_data.items() if form_data else {})
            if v is not None
        }

        # Required: customer name
        customer_name = form_data.get('customer_name')
        if not customer_name:
            return False, {"error": "Customer name is required"}

        # Find or create customer
        customer = Customer.query.filter_by(name=customer_name).first()
        if not customer:
            # if a customer_fee is provided, try to parse it
            fee = None
            if 'customer_fee' in form_data and form_data.get('customer_fee') not in (None, ""):
                try:
                    fee_str = str(form_data.get('customer_fee')).replace(',', '.').strip()
                    fee = float(fee_str)
                except (ValueError, TypeError):
                    fee = None
            customer = Customer(name=customer_name, fee=fee if fee is not None else 0.0)
            db.session.add(customer)
            db.session.flush()  # ensure ID assigned

        # Form number allocation (with optional requested start)
        requested_start = None
        if form_data.get('start_form_number'):
            try:
                requested_start = int(str(form_data.get('start_form_number')).strip())
                if requested_start <= 0:
                    requested_start = None
            except ValueError:
                requested_start = None

        try:
            form_number = allocate_form_number_for_year(start_from=requested_start)
        except Exception as e:
            db.session.rollback()
            print("Form number allocation failed:", str(e))
            return False, {"error": "Failed to allocate form number due to DB error."}

        # Parse dates (delivery and exits)
        delivery_date = parse_date_input(form_data.get('delivery_date'))
        if form_data.get('delivery_date') and delivery_date is None:
            return False, {"error": "Invalid delivery date format. Use YYYY-MM-DD or YYYY/MM/DD format"}

        exit_from_office_date = parse_date_input(form_data.get('exit_from_office_date'))
        if form_data.get('exit_from_office_date') and exit_from_office_date is None:
            return False, {"error": "Invalid exit from office date format. Use YYYY-MM-DD or YYYY/MM/DD format"}

        exit_from_factory_date = parse_date_input(form_data.get('exit_from_factory_date'))
        if form_data.get('exit_from_factory_date') and exit_from_factory_date is None:
            return False, {"error": "Invalid exit from factory date format. Use YYYY-MM-DD or YYYY/MM/DD format"}

        # Convert numeric fields robustly
        try:
            width = float(form_data['width']) if form_data.get('width') not in (None, "") else None
            height = float(form_data['height']) if form_data.get('height') not in (None, "") else None
            quantity = int(form_data['quantity']) if form_data.get('quantity') not in (None, "") else None
            total_length_meters = float(form_data['total_length_meters']) if form_data.get('total_length_meters') not in (None, "") else None
            peak_quantity = int(form_data['peak_quantity']) if form_data.get('peak_quantity') not in (None, "") else None
        except (ValueError, TypeError) as e:
            return False, {"error": f"Invalid numeric value: {str(e)}"}

        # Compute sketch_name with prefix based on width
        raw_sketch = form_data.get('sketch_name') or ""
        base_sketch = strip_known_prefixes(raw_sketch)
        prefix = get_sketch_prefix_for_width(width)
        sketch_name_final = _apply_prefix_to_sketch(prefix, base_sketch)

        # Create new order
        new_order = Order(
            form_number=form_number,
            customer_id=customer.id,
            fabric_density=form_data.get('fabric_density'),
            fabric_cut=form_data.get('fabric_cut'),
            width=width,
            height=height,
            quantity=quantity,
            total_length_meters=total_length_meters,
            peak_quantity=peak_quantity,
            delivery_date=delivery_date,
            exit_from_office_date=exit_from_office_date,
            exit_from_factory_date=exit_from_factory_date,
            sketch_name=sketch_name_final,
            file_name=form_data.get('file_name'),
            design_specification=form_data.get('design_specification'),
            office_notes=form_data.get('office_notes'),
            factory_notes=form_data.get('factory_notes'),
            customer_note_to_office=form_data.get('customer_note_to_office'),
            fusing_type=form_data.get('fusing_type'),
            lamination_type=form_data.get('lamination_type'),
            cut_type=form_data.get('cut_type'),
            label_type=form_data.get('label_type'),
            created_by=current_user.id,
            status=form_data.get('status', 'Pending'),
        )

        db.session.add(new_order)
        db.session.flush()

        # --- Save Order Values --- #
        values = form_data.get('values[]') or form_data.get('values')
        if values:
            if isinstance(values, str):
                values = [values]
            for idx, value in enumerate(values, 1):
                if value is not None and str(value).strip() != "":
                    db.session.add(OrderValue(order_id=new_order.id, value_index=idx, value=value))

        # --- Save Order Files robustly --- #
        if hasattr(form_data, 'getlist'):
            file_display_names = form_data.getlist('file_display_names[]')
            file_names = form_data.getlist('order_files[]')
        else:
            file_display_names = form_data.get('file_display_names[]') or []
            file_names = form_data.get('order_files[]') or []
            if isinstance(file_display_names, str):
                file_display_names = [file_display_names]
            if isinstance(file_names, str):
                file_names = [file_names]

        max_len = max(len(file_display_names), len(file_names))
        file_display_names += [""] * (max_len - len(file_display_names))
        file_names += [""] * (max_len - len(file_names))
        for display_name, file_name in zip(file_display_names, file_names):
            if display_name or file_name:
                order_file = OrderFile(
                    order_id=new_order.id,
                    display_name=display_name,
                    file_name=file_name,
                    uploaded_by=current_user.id
                )
                db.session.add(order_file)

        db.session.commit()

        # Handle image uploads if provided
        if files and hasattr(files, 'getlist') and 'images' in files:
            image_files = files.getlist('images')
            for file in image_files:
                if file and getattr(file, 'filename', None):
                    success, response = upload_order_image(new_order.id, file)
                    if not success:
                        # Log warning but continue
                        print(f"Warning: Failed to upload image {file.filename}: {response.get('error')}")

        return True, {
            "message": "Order created successfully",
            "order": new_order.to_dict()
        }

    except Exception as e:
        db.session.rollback()
        error_traceback = traceback.format_exc()
        print("Error creating order:")
        print(error_traceback)
        return False, {"error": f"Failed to create order: {str(e)}\nTraceback: {error_traceback}"}


def get_orders(page: int = 1, per_page: int = 10, search: str = None, status: str = None) -> Tuple[bool, Dict[str, Any]]:
    """
    Get all orders with their details, with pagination and optional search/status filters.
    Returns a tuple of (success, response) where response contains either the orders list or an error message.
    """
    try:
        query = Order.query.join(Customer)  # ✅ join with customer so we can filter by name

        # Apply search filter
        if search:
            query = query.filter(
                db.or_(
                    Customer.name.ilike(f'%{search}%'),                 # ✅ search customer name
                    db.cast(Order.form_number, db.String).ilike(f'%{search}%')  # ✅ search form number
                )
            )
        
        # Apply status filter
        if status and status.lower() != 'all':
            query = query.filter(db.func.lower(Order.status) == status.lower())

        # Order by form_number descending (highest to lowest)
        query = query.order_by(Order.form_number.desc())
        
        # Paginate the results
        pagination = query.paginate(page=page, per_page=per_page, error_out=False)
        
        # Convert orders to list of dictionaries
        orders_list = [order.to_dict() for order in pagination.items]
        
        return True, {
            "message": "Orders retrieved successfully",
            "orders": orders_list,
            "total": pagination.total,
            "pagination":pagination
        }
        
    except Exception as e:
        print(f"Error retrieving orders: {str(e)}")
        return False, {"error": "Failed to retrieve orders"}
    
def get_order_by_id(order_id: int) -> Tuple[bool, Dict[str, Any]]:
    """
    Get a specific order by its ID.
    """
    try:
        order = Order.query.get(order_id)
        if not order:
            return False, {"error": "Order not found"}
        
        return True, {
            "message": "Order retrieved successfully",
            "order": order.to_dict()
        }
        
    except Exception as e:
        print(f"Error retrieving order {order_id}: {str(e)}")
        return False, {"error": f"Failed to retrieve order: {str(e)}"}
    
def delete_order_by_id(order_id: int) -> Tuple[bool , Dict[str, Any]]:
    """
    Delete an order by its ID.
    """
    try:
        order = Order.query.get(order_id)
        if not order:
            return False, {"error": "Order not found"}
        
        db.session.delete(order)
        db.session.commit()
        
        return True, {
            "message": "سفارش با موفقیت حذف شد",
        }
        
    except Exception as e:
        db.session.rollback()
        print(f"Error deleting order {order_id}: {str(e)}")
        return False, {"error": f"Failed to delete order: {str(e)}"}

def update_order_id(order_id: int, form_data: Dict[str, Any], files=None) -> Tuple[bool, Dict[str, Any]]:
    """
    Update an existing order with provided form_data.
    Handles customer_name -> customer_id resolution and updates customer.fee if provided.
    Also applies sketch_name prefixing based on width.
    """
    try:
        order = Order.query.get(order_id)
        if not order:
            print(f"Order {order_id} not found")
            return False, {"error": "Order not found"}

        # --- Handle customer_name separately (create if not exists) --- #
        customer = None
        if form_data.get('customer_name'):
            customer_name = str(form_data['customer_name']).strip()
            customer = Customer.query.filter_by(name=customer_name).first()
            if not customer:
                customer = Customer(name=customer_name)
                db.session.add(customer)
                db.session.flush()
            order.customer_id = customer.id
        else:
            customer = order.customer

        # --- Update customer fee if provided in form_data --- #
        if 'customer_fee' in form_data:
            fee_raw = form_data.get('customer_fee')
            if fee_raw is not None and fee_raw != "":
                try:
                    fee_str = str(fee_raw).replace(',', '.').strip()
                    fee_value = float(fee_str)
                except (ValueError, TypeError):
                    return False, {"error": f"Invalid customer_fee value: {fee_raw}"}
                if not customer:
                    if order.customer_id:
                        customer = Customer.query.get(order.customer_id)
                    else:
                        customer = Customer(name="Unknown Customer")
                        db.session.add(customer)
                        db.session.flush()
                        order.customer_id = customer.id
                customer.fee = fee_value

        # --- Update other order fields (skip customer_name & customer_fee) --- #
        for key, value in (form_data.items() if form_data else {}):
            if key in ('customer_name', 'customer_fee'):
                continue
            if hasattr(order, key):
                if key == 'created_at' and value:
                    parsed_date = parse_date_input(value)
                    if parsed_date is None:
                        return False, {"error": f"Invalid created_at format: {value}. Use YYYY-MM-DD or YYYY/MM/DD"}
                    value = datetime.combine(parsed_date, datetime.min.time())
                elif key == 'delivery_date' and value:
                    parsed_date = parse_date_input(value)
                    if parsed_date is None:
                        return False, {"error": f"Invalid delivery_date format: {value}. Use YYYY-MM-DD or YYYY/MM/DD"}
                    value = parsed_date
                elif key in ['exit_from_office_date', 'exit_from_factory_date'] and value:
                    parsed_date = parse_date_input(value)
                    if parsed_date is None:
                        return False, {"error": f"Invalid {key} format: {value}. Use YYYY-MM-DD or YYYY/MM/DD"}
                    value = parsed_date
                else:
                    # convert numeric strings to proper types for width/height/quantity etc if necessary
                    if key in ('width', 'height', 'total_length_meters', 'fabric_cut', 'fabric_density') and value not in (None, ""):
                        try:
                            value = float(str(value).replace(',', '.').strip())
                        except (ValueError, TypeError):
                            # leave raw value; DB will error if invalid, or we can return error
                            return False, {"error": f"Invalid numeric value for {key}: {value}"}
                    if key in ('quantity', 'peak_quantity') and value not in (None, ""):
                        try:
                            value = int(str(value).strip())
                        except (ValueError, TypeError):
                            return False, {"error": f"Invalid integer value for {key}: {value}"}
                setattr(order, key, value)
            else:
                print(f"Field {key} not found in Order model")

        # --- PATCH-like update for Order Values (edit-values / values) --- #
        values = (
            form_data.get('edit-values[]')
            or form_data.get('edit-values')
            or form_data.get('values[]')
            or form_data.get('values')
        )
        if values:
            if isinstance(values, str):
                values = [values]
            values = list(values)
            while len(values) < 8:
                values.append("")
            for idx in range(1, 9):
                value = values[idx - 1] if idx - 1 < len(values) else ""
                order_value = OrderValue.query.filter_by(order_id=order.id, value_index=idx).first()
                if order_value:
                    order_value.value = value
                else:
                    db.session.add(OrderValue(order_id=order.id, value_index=idx, value=value))

        # --- Normalize file lists (edit-file_display_names / edit-file_names / existing_file_ids) --- #
        if hasattr(form_data, 'getlist'):
            file_display_names = form_data.getlist('edit-file_display_names[]')
            file_names = form_data.getlist('edit-file_names[]')
            existing_file_ids = form_data.getlist('existing_file_ids[]')
        else:
            file_display_names = form_data.get('edit-file_display_names[]') or []
            file_names = form_data.get('edit-file_names[]') or []
            existing_file_ids = form_data.get('existing_file_ids[]') or []
            if isinstance(file_display_names, str):
                file_display_names = [file_display_names]
            if isinstance(file_names, str):
                file_names = [file_names]
            if isinstance(existing_file_ids, str):
                existing_file_ids = [existing_file_ids]

        max_len = max(len(file_display_names), len(file_names), len(existing_file_ids))
        file_display_names += [""] * (max_len - len(file_display_names))
        file_names += [""] * (max_len - len(file_names))
        existing_file_ids += [""] * (max_len - len(existing_file_ids))

        # --- Update or Add OrderFiles --- #
        for file_id, display_name, file_name in zip_longest(existing_file_ids, file_display_names, file_names, fillvalue=""):
            if file_id:
                try:
                    order_file = OrderFile.query.get(int(file_id))
                except (ValueError, TypeError):
                    order_file = None
                if order_file:
                    order_file.display_name = display_name
                    order_file.file_name = file_name
            else:
                if display_name or file_name:
                    new_file = OrderFile(
                        order_id=order.id,
                        display_name=display_name,
                        file_name=file_name,
                        uploaded_by=current_user.id
                    )
                    db.session.add(new_file)

        # --- Apply sketch_name prefixing based on effective width --- #
        # Determine width to use: prefer incoming form value, else existing order.width
        effective_width = None
        if 'width' in form_data and form_data.get('width') not in (None, ""):
            try:
                effective_width = float(str(form_data.get('width')).replace(',', '.').strip())
            except (ValueError, TypeError):
                effective_width = None
        else:
            effective_width = getattr(order, 'width', None)

        if 'sketch_name' in form_data:
            incoming_sketch_raw = (form_data.get('sketch_name') or "").strip()
        else:
            incoming_sketch_raw = (order.sketch_name or "").strip()

        # Strip any existing known prefix(es) before applying the new one
        incoming_sketch_base = strip_known_prefixes(incoming_sketch_raw)

        prefix = get_sketch_prefix_for_width(effective_width)
        order.sketch_name = _apply_prefix_to_sketch(prefix, incoming_sketch_base)
        order.updated_at = datetime.utcnow()
        db.session.commit()
        db.session.refresh(order)

        return True, {
            "message": "Order updated successfully",
            "order": order.to_dict()
        }

    except Exception as e:
        db.session.rollback()
        print(f"Error updating order {order_id}: {str(e)}")
        print("Full error details:", traceback.format_exc())
        return False, {"error": f"Failed to update order: {str(e)}"}

def duplicate_order(order_id):
    """
    Duplicate an existing order with a new ID, including its OrderFiles, OrderValues, and OrderImages.
    """
    try:
        # Step 1: Fetch the original order
        success, response = get_order_by_id(order_id)
        if not success or not response.get('order'):
            return False, {"error": "Order not found"}

        original_order = response['order']
        form_number = _get_next_form_number_for_year()  # Generate new form number

        # Step 2: Create new order with copied values
        new_order_data = {
            'form_number': form_number,
            'customer_id': original_order.get('customer_id'),  # ✅ use customer_id
            'customer_name': original_order.get('customer_name'),  # ensure add_order passes customer validation
            'fabric_density': original_order.get('fabric_density'),
            'fabric_cut': original_order.get('fabric_cut'),
            'width': original_order.get('width'),
            'height': original_order.get('height'),
            'quantity': original_order.get('quantity'),
            'total_length_meters': original_order.get('total_length_meters'),
            'peak_quantity':original_order.get('peak_quantity'),
            'fusing_type': original_order.get('fusing_type'),
            'lamination_type': original_order.get('lamination_type'),
            'cut_type': original_order.get('cut_type'),
            'label_type': original_order.get('label_type'),
            'delivery_date': original_order.get('delivery_date'),
            'exit_from_office_date': original_order.get('exit_from_office_date'),
            'exit_from_factory_date': original_order.get('exit_from_factory_date'),
            'sketch_name': original_order.get('sketch_name'),
            'file_name': original_order.get('file_name'),
            'status': 'Pending',
            'design_specification': original_order.get('design_specification'),
            'office_notes': "برابر تولیدات قبل",
            'factory_notes': original_order.get('factory_notes'),
            'customer_note_to_office': original_order.get('customer_note_to_office')
        }

        success, response = add_order(new_order_data)
        if not success:
            return False, response

        new_order = Order.query.get(response['order']['id'])
        orig_order_obj = Order.query.get(order_id)

        # Step 3: Copy OrderValues
        for val in orig_order_obj.values:
            db.session.add(OrderValue(
                order_id=new_order.id,
                value_index=val.value_index,
                value=val.value
            ))

        # Step 4: Copy OrderFiles
        for file in orig_order_obj.files:
            db.session.add(OrderFile(
                order_id=new_order.id,
                file_name=file.file_name,
                display_name=file.display_name,
                uploaded_by=file.uploaded_by
            ))

        # Step 5: Copy OrderImages
        for image in orig_order_obj.images:
            db.session.add(OrderImage(
                order_id=new_order.id,
                filename=image.filename,
                original_filename=image.original_filename,
                file_path=image.file_path,
                file_size=image.file_size,
                mime_type=image.mime_type,
                uploaded_by=image.uploaded_by
            ))

        db.session.commit()
        return True, {"order": new_order.to_dict()}

    except Exception as e:
        db.session.rollback()
        print(f"Error in duplicate_order: {str(e)}")
        return False, {"error": "An error occurred while duplicating the order"}

def generate_excel_report(search: str = None, status: str = None) -> Tuple[bool, Dict[str, Any]]:
    """
    Generate a Persian Excel report of orders with RTL layout and formatted columns.
    Returns (success, response) where response contains either the Excel file buffer or an error message.
    """
    try:
        query = Order.query

        # Apply filters
        if search:
            query = query.filter(
                db.or_(
                    Order.customer_name.ilike(f"%{search}%"),
                    db.cast(Order.form_number, db.String).ilike(f"%{search}%")
                )
            )

        if status and status.lower() != 'all':
            query = query.filter(db.func.lower(Order.status) == status.lower())

        orders = query.order_by(Order.created_at.desc()).all()

        wb = Workbook()
        ws = wb.active
        ws.title = "گزارش سفارشات"

        # Set worksheet direction to RTL
        ws.sheet_view.rightToLeft = True

        # Persian headers
        headers = [
            "شماره فرم",
            "نام مشتری",
            "نام طرح",
            "تعداد",
            "متر کل",
            "خروج از دفتر",
            "خروج از کارخانه",
            "تاریخ تحویل",
            "آخرین بروزرسانی"
        ]

        column_widths = {
            'A': 15,
            'B': 25,
            'C': 25,
            'D': 10,
            'E': 12,
            'F': 18,
            'G': 18,
            'H': 18,
            'I': 20
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        header_font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        data_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        data_font = Font(name="B Kamran", size=11)

        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Add order data
        for row_idx, order in enumerate(orders, 2):
            customer_name = order.customer.name if getattr(order, 'customer', None) else None
            values = [
                order.form_number,
                customer_name,
                order.sketch_name,
                order.quantity,
                order.total_length_meters,
                order.peak_quantity,
                order.exit_from_office_date,
                order.exit_from_factory_date,
                order.delivery_date,
                order.updated_at
            ]

            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = data_font
                cell.alignment = data_alignment

                # Format dates
                if isinstance(value, datetime):
                    cell.number_format = "yyyy-mm-dd hh:mm:ss"
                elif hasattr(value, 'strftime'):
                    cell.number_format = "yyyy-mm-dd"

                # Zebra row fill
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        # Borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in ws.iter_rows(min_row=1, max_row=len(orders)+1, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border

        # Freeze header
        ws.freeze_panes = 'A2'

        # Save to memory
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"گزارش_سفارشات_{timestamp}.xlsx"

        return True, {
            "excel_file_buffer": excel_file,
            "file_name": filename
        }

    except Exception as e:
        print(f"❌ Error generating Excel report: {str(e)}")
        traceback.print_exc()
        return False, {"error": "خطا در تولید فایل اکسل."}

def _allowed_file(filename: str) -> bool:
    """Check if the file extension is allowed"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def _ensure_upload_folder():
    """Ensure the upload folder exists"""
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)

def upload_order_image(order_id: int, file) -> Tuple[bool, Dict[str, Any]]:
    """
    Upload an image for an order.
    Returns (success, response) where response contains either the image data or an error message.
    """
    try:
        # Check if order exists
        order = Order.query.get(order_id)
        if not order:
            return False, {"error": "Order not found"}

        # Validate file
        if not file or not file.filename:
            return False, {"error": "No file provided"}
        
        # Check if file has an extension
        original_filename = secure_filename(file.filename)
        if '.' not in original_filename:
            return False, {"error": "File must have an extension"}
        
        if not _allowed_file(original_filename):
            return False, {"error": f"File type not allowed. Allowed types: {', '.join(ALLOWED_EXTENSIONS)}"}
        
        if file.content_length and file.content_length > MAX_FILE_SIZE:
            return False, {"error": f"File too large. Maximum size is {MAX_FILE_SIZE/1024/1024}MB"}

        # Ensure upload folder exists
        _ensure_upload_folder()

        # Generate unique filename
        file_ext = original_filename.rsplit('.', 1)[-1].lower()
        unique_filename = f"{uuid.uuid4().hex}.{file_ext}"
        
        # Save file
        file_path = os.path.join(UPLOAD_FOLDER, unique_filename)
        file.save(file_path)
        
        # Get file size
        file_size = os.path.getsize(file_path)
        
        # Create image record
        image = OrderImage(
            order_id=order_id,
            filename=unique_filename,
            original_filename=original_filename,
            file_path=file_path,
            file_size=file_size,
            mime_type=file.content_type,
            uploaded_by=current_user.id
        )
        
        db.session.add(image)
        db.session.commit()
        
        return True, {
            "message": "Image uploaded successfully",
            "image": image.to_dict()
        }
        
    except Exception as e:
        db.session.rollback()
        print(f"Error uploading image: {str(e)}")
        print(traceback.format_exc())
        return False, {"error": f"Failed to upload image: {str(e)}"}

def delete_order_image(image_id: int) -> Tuple[bool, Dict[str, Any]]:
    """
    Delete an order image.
    Returns (success, response) where response contains either a success message or an error.
    """
    try:
        image = OrderImage.query.get(image_id)
        if not image:
            return False, {"error": "Image not found"}
        
        # Delete file from filesystem
        try:
            if os.path.exists(image.file_path):
                os.remove(image.file_path)
        except Exception as e:
            print(f"Error deleting file: {str(e)}")
        
        # Delete from database
        db.session.delete(image)
        db.session.commit()
        
        return True, {"message": "Image deleted successfully"}
        
    except Exception as e:
        db.session.rollback()
        print(f"Error deleting image: {str(e)}")
        return False, {"error": f"Failed to delete image: {str(e)}"}

def get_order_images(order_id: int) -> Tuple[bool, Dict[str, Any]]:
    """
    Get all images for an order.
    Returns (success, response) where response contains either the images list or an error.
    """
    try:
        order = Order.query.get(order_id)
        if not order:
            return False, {"error": "Order not found"}
        
        images = [image.to_dict() for image in order.images]
        return True, {
            "message": "Images retrieved successfully",
            "images": images
        }
        
    except Exception as e:
        print(f"Error retrieving images: {str(e)}")
        return False, {"error": f"Failed to retrieve images: {str(e)}"}

def _float_eq(a: float, b: float, tol: float = 1e-3) -> bool:
    try:
        return abs(float(a) - float(b)) <= tol
    except Exception:
        return False

def get_sketch_prefix_for_width(width: Optional[float]) -> Optional[str]:
    """
    Return Persian prefix for the provided numeric width according to the mapping:
      1.1, 1.25, 1.32, 1.42 -> "باریک"
      1.53, 1.67, 1.8, 2, 2.2, 2.5, 2.85 -> "اتیکت"
      3.3, 4, 5, 6.67 -> "کتی"
    Returns None when no matching prefix found.
    """
    if width is None:
        return None
    try:
        w = float(width)
    except (TypeError, ValueError):
        return None

    barik = {1.1, 1.25, 1.32, 1.42}
    etiket = {1.53, 1.67, 1.8, 2.0, 2.2, 2.5, 2.85}
    kati = {3.3, 4.0, 5.0, 6.67 , 6 , 7 , 8 , 10}

    for val in barik:
        if _float_eq(w, val):
            return "باریک"
    for val in etiket:
        if _float_eq(w, val):
            return "اتیکت"
    for val in kati:
        if _float_eq(w, val):
            return "کتی"

    return None

def _apply_prefix_to_sketch(prefix: Optional[str], raw_sketch: Optional[str]) -> Optional[str]:
    """
    Combine prefix and raw_sketch idempotently.
    If prefix is None -> return stripped raw_sketch or None.
    """
    raw_sketch_norm = (raw_sketch or "").strip()
    if not prefix:
        return raw_sketch_norm or None
    # if already prefixed with the same string, return as-is
    if raw_sketch_norm.startswith(prefix):
        return raw_sketch_norm or prefix
    # if sketch empty, return prefix alone
    if raw_sketch_norm == "":
        return prefix
    return f"{prefix} {raw_sketch_norm}"
KNOWN_PREFIXES = ["باریک", "اتیکت", "کتی"]

def strip_known_prefixes(text: Optional[str]) -> str:
    """
    Remove any known prefix(es) from the start of `text`.
    Example: "باریک اتیکت nike" -> "nike"
    Returns stripped text (empty string if nothing left).
    """
    if not text:
        return ""
    s = text.strip()
    # loop to remove multiple prefixes if present
    removed = True
    while removed:
        removed = False
        for p in KNOWN_PREFIXES:
            # if it starts with prefix + space OR exactly prefix, remove it
            if s == p:
                s = ""
                removed = True
                break
            if s.startswith(p + " "):
                s = s[len(p):].strip()
                removed = True
                break
    return s
