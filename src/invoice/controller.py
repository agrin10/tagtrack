from flask import render_template, request, jsonify, redirect, url_for
from src.invoice.models import Payment , InvoiceDraft
from src import db
from datetime import datetime
from typing import Tuple, Dict, Any
import uuid
from src.order.models import Order
import io
from flask import send_file
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
import openpyxl
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from bidi.algorithm import get_display
import arabic_reshaper
import os
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from decimal import Decimal ,  ROUND_HALF_UP
from datetime import datetime, date, timezone
from src.production.models import ProductionStepEnum , ProductionStepLog
# Register Persian font
FONT_PATH = os.path.abspath('src/static/fonts/Vazir.ttf')

pdfmetrics.registerFont(TTFont('Vazir', FONT_PATH))

# Function to reshape and display Persian text
def persian(text):
    if not text:
        return ""
    reshaped_text = arabic_reshaper.reshape(text)
    return get_display(reshaped_text)

def invoice_list(page: int = 1, per_page: int = 10, search: str = None, status: str = None) -> Tuple[bool, Dict[str, Any]]:
    try:
        query = Payment.query

        if search:
            query = query.filter(
                db.or_(
                    Payment.customer_name.ilike(f'%{search}%'),
                    db.cast(Payment.form_number, db.String).ilike(f'%{search}%')
                )
            )

        if status and status.lower() != 'all':
            query = query.filter(db.func.lower(Payment.status) == status.lower())

        query = query.order_by(Payment.created_at.desc())
        pagination = query.paginate(page=page, per_page=per_page, error_out=False)

        invoices_list = [payment.to_dict() for payment in pagination.items]

        return True, {
            "message": "Orders retrieved successfully",
            "payments": invoices_list,
            "total": pagination.total,
            "pagination": pagination
        }
        

    except Exception as e:
        print(f"Error retrieving orders: {str(e)}")
        return False, {"error": f"Failed to retrieve orders: {e}"}

def get_invoice_for_order(order_id: int) -> Tuple[bool, Dict[str, Any]]:
    """
    Get invoice data for a specific order to populate factory processing modal.
    """
    try:
        invoice = Payment.query.filter_by(order_id=order_id).first()
        if not invoice:
            return False, {"message": "No invoice found for this order"}
        
        return True, {
            "message": "Invoice data retrieved successfully",
            "invoice": invoice.to_dict()
        }
        
    except Exception as e:
        print(f"Error retrieving invoice for order {order_id}: {str(e)}")
        return False, {"error": f"Failed to retrieve invoice: {str(e)}"}

def view_invoice(invoice_id: int) -> Tuple[bool, Dict[str, Any]]:
    """
    Get a specific invoice by its ID.
    """
    try:
        payment = Payment.query.get(invoice_id)
        if not payment:
            return False, {"error": "invoice not found"}
        
        return True, {
            "message": "invoice retrieved successfully",
            "invoice": payment.to_dict()
        }
        
    except Exception as e:
        print(f"Error retrieving invoice {invoice_id}: {str(e)}")
        return False, {"error": f"Failed to retrieve invoice: {str(e)}"}


def download_invoice(invoice_id, file_type):
    success, data = view_invoice(invoice_id)
    if not success or not data:
        return False, None

    invoice = data["invoice"]

    if file_type == 'pdf':
        buffer = io.BytesIO()
        p = canvas.Canvas(buffer, pagesize=A4)
        width, height = A4

        # Header
        p.setFont("Vazir", 22)
        p.setFillColor(colors.darkblue)
        p.drawRightString(width - 50, height - 60, persian("فاکتور"))

        p.setFont("Vazir", 12)
        p.setFillColor(colors.black)
        p.drawRightString(width - 50, height - 90, persian(f"تاریخ: {invoice['issue_date'][:10]}"))
        p.drawRightString(width - 50, height - 110, persian(f"شماره فاکتور: {invoice['invoice_number']}"))
        p.drawRightString(width - 50, height - 130, persian(f"شماره فرم: {invoice.get('form_number', '---')}"))

        p.setStrokeColor(colors.lightgrey)
        p.setLineWidth(0.5)
        p.line(50, height - 140, width - 50, height - 140)

        # Content
        y_start = height - 180
        line_height = 20
        labels_left = [
            (persian("قیمت واحد"), f"{invoice['unit_price']}"),
            (persian("تعداد تولیدی"), f"{invoice['quantity']}"),
            (persian("تعداد پیک"), f"{invoice['peak_quantity']}"),
            (persian("ردیف"), f"{invoice.get('row_number', '---')}"),
        ]
        labels_right = [
            (persian("هزینه برش"), f"{invoice['cutting_cost']}"),
            (persian("قیمت کل"), f"{invoice['total_price']}"),
            (persian("وضعیت"), persian(invoice['status'])),
            (persian("یادداشت"), persian(invoice['notes'] or "---")),
        ]

        p.setFont("Vazir", 11)
        for i, (label, value) in enumerate(labels_left):
            y = y_start - (i * line_height)
            p.drawRightString(width - 350, y, label + ":")
            p.setFont("Vazir", 10)
            p.drawRightString(width - 450, y, str(value))
            p.setFont("Vazir", 11)

        for i, (label, value) in enumerate(labels_right):
            y = y_start - (i * line_height)
            p.drawRightString(width - 80, y, label + ":")
            p.setFont("Vazir", 10)
            p.drawRightString(width - 180, y, str(value))
            p.setFont("Vazir", 11)

        # Footer
        p.setStrokeColor(colors.lightgrey)
        p.line(50, 80, width - 50, 80)
        p.setFont("Vazir", 9)
        p.setFillColor(colors.grey)
        p.drawRightString(width - 50, 65, persian("با تشکر از خرید شما"))
        p.drawString(50, 65, persian("تولید شده توسط سیستم فاکتور AM"))

        p.showPage()
        p.save()
        buffer.seek(0)

        return True, send_file(
            buffer,
            as_attachment=True,
            download_name=f"{invoice['invoice_number']}.pdf",
            mimetype='application/pdf'
        )

    elif file_type == 'excel':
        FIELD_TRANSLATIONS = {
            "invoice_number": "شماره فاکتور",
            "issue_date": "تاریخ صدور",
            "form_number": "شماره فرم",
            "unit_price": "قیمت واحد",
            "quantity": "تعداد تولیدی",
            "peak_quantity": "تعداد پیک",
            "row_number": "ردیف",
            "cutting_cost": "هزینه برش",
            "total_price": "قیمت کل",
            "status": "وضعیت",
            "notes": "یادداشت",
        }

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "فاکتور"

        # Styling objects
        rtl_alignment = Alignment(horizontal='right', readingOrder=2)
        header_fill = PatternFill(start_color="0d6efd", end_color="0d6efd", fill_type="solid")
        odd_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        font_bold = Font(name='Tahoma', bold=True)
        font_normal = Font(name='Tahoma')

        # Write headers
        headers = [FIELD_TRANSLATIONS[key] for key in FIELD_TRANSLATIONS]
        ws.append(headers)
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col)
            cell.alignment = rtl_alignment
            cell.font = font_bold
            cell.fill = header_fill

        # Write values
        values = [invoice.get(key, "---") for key in FIELD_TRANSLATIONS]
        ws.append(values)
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=2, column=col)
            cell.alignment = rtl_alignment
            cell.font = font_normal
            cell.fill = odd_fill  # light gray background

        # Adjust column widths
        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max(12, max_length + 3)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return True, send_file(
            buffer,
            as_attachment=True,
            download_name=f"{invoice['invoice_number']}.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    return False, None
def export_all():
    """
    Export all invoices to an Excel file.
    """
    try:
        FIELD_TRANSLATIONS = {
            "invoice_number": "شماره فاکتور",
            "issue_date": "تاریخ صدور",
            "form_number": "شماره فرم",
            "unit_price": "قیمت واحد",
            "quantity": "تعداد تولیدی",
            "peak_quantity": "تعداد پیک",
            "row_number": "ردیف",
            "cutting_cost": "هزینه برش",
            "total_price": "قیمت کل",
            "status": "وضعیت",
            "notes": "یادداشت",
        }

        invoices = Payment.query.order_by(Payment.created_at.desc()).all()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "فاکتورها"
        default_fill = PatternFill(fill_type=None)  # for rows that don't use odd_fill

        # Styling objects
        rtl_alignment = Alignment(horizontal='right', readingOrder=2)
        header_fill = PatternFill(start_color="0d6efd", end_color="0d6efd", fill_type="solid")
        odd_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        font_bold = Font(name='Tahoma', bold=True)
        font_normal = Font(name='Tahoma')

        # Write headers
        headers = [FIELD_TRANSLATIONS[key] for key in FIELD_TRANSLATIONS]
        ws.append(headers)
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col)
            cell.alignment = rtl_alignment
            cell.font = font_bold
            cell.fill = header_fill

        # Write invoice rows
        for idx, invoice in enumerate(invoices, start=2):
            inv_dict = invoice.to_dict()
            values = [inv_dict.get(key, "---") for key in FIELD_TRANSLATIONS]
            ws.append(values)
            for col, value in enumerate(values, start=1):
                cell = ws.cell(row=idx, column=col)
                cell.alignment = rtl_alignment
                cell.font = font_normal
                cell.fill = odd_fill if idx % 2 == 0 else default_fill
        # Adjust column widths
        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max(12, max_length + 3)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return send_file(
            buffer,
            as_attachment=True,
            download_name="all_invoices.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        ) , 200
    except Exception as e:
        return jsonify({"error": f"Failed to export invoices: {e}"}), 500
    
def _to_decimal(value, name="value"):
    try:
        if value is None or (isinstance(value, str) and value.strip() == ""):
            return Decimal('0')
        return Decimal(str(value))
    except Exception as e:
        raise ValueError(f"Invalid numeric for {name}: {value} ({e})")

def save_invoice_from_factory(order_id: int, invoice_data: Dict[str, Any], user_id: int) -> Dict[str, Any]:
    """
    Robust save/update invoice using invoice_data (prefer) and fallbacks to order/customer.
    Calculation formula used:
      price = peak_quantity * peak_width * fee
      unit_price = price + cutting_cost
      total_price = unit_price * number_of_cuts   # matches your Persian formula
    Uses Decimal for arithmetic and returns formatted numbers.
    """
    try:
        order = Order.query.get(order_id)
        if not order:
            return {"success": False, "message": "Order not found"}

        existing_invoice = Payment.query.filter_by(order_id=order_id).first()
        existing_draft = InvoiceDraft.query.filter_by(order_id=order_id).first()

        # Extract / prefer invoice_data, fallback to order/customer
        qty_raw = invoice_data.get('quantity', invoice_data.get('Quantity', None))
        cutting_cost_raw = invoice_data.get('cutting_cost', invoice_data.get('cuttingCost', 0.0))
        cuts_raw = invoice_data.get('number_of_cuts', invoice_data.get('number_of_cuts', None))
        peak_quantity_raw = invoice_data.get('peak_quantity', invoice_data.get('peakQuantity', None))
        peak_width_raw = invoice_data.get('peak_width', None)
        if peak_width_raw is None:
            peak_width_raw = invoice_data.get('peakWidth', None)
        fee_raw = invoice_data.get('Fee', invoice_data.get('fee', None))  # accept Fee or fee
        row_number = invoice_data.get('row_number', None)
        notes = invoice_data.get('notes', 'Generated from factory processing')

        # sensible fallbacks from order / customer
        if peak_width_raw is None:
            peak_width_raw = getattr(order, 'width', None)
        if fee_raw is None:
            fee_raw = getattr(order.customer, 'fee', None) if getattr(order, 'customer', None) else None
        if qty_raw is None:
            qty_raw = getattr(order, 'quantity', None)

        # If number_of_cuts not provided, fallback to quantity (business choice)
        if cuts_raw is None:
            cuts_raw = qty_raw

        # Convert to Decimal safely
        try:
            quantity = _to_decimal(qty_raw, "quantity")
            cutting_cost = _to_decimal(cutting_cost_raw, "cutting_cost")
            number_of_cuts = _to_decimal(cuts_raw, "number_of_cuts")
            peak_quantity = _to_decimal(peak_quantity_raw, "peak_quantity")
            peak_width = _to_decimal(peak_width_raw, "peak_width")
            fee = _to_decimal(fee_raw, "fee")
        except ValueError as ve:
            return {"success": False, "message": str(ve)}

        # Validation: required positives (adjust rules if you want different)
        if peak_quantity <= 0 or peak_width <= 0 or fee <= 0 or number_of_cuts <= 0:
            return {"success": False, "message": "peak_quantity, peak_width, fee and number_of_cuts must be positive."}
        press_cost = _compute_press_cost_one_time(order_id)  # Decimal('350') or Decimal('0')
        lamination_cost = press_cost
        # Calculation with Decimal and rounding to 2 decimal places
        price = (peak_quantity * peak_width * fee).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        services = (_to_decimal(cutting_cost, "cutting_cost") + press_cost).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        unit_price = (price + services).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        total_price = (unit_price * number_of_cuts).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

        # Update existing invoice
        if existing_invoice:
            existing_invoice.unit_price = float(unit_price)
            existing_invoice.quantity = int(quantity) if quantity == quantity.to_integral() else float(quantity)
            existing_invoice.cutting_cost = float(cutting_cost)
            existing_invoice.number_of_cuts = int(number_of_cuts) if number_of_cuts == number_of_cuts.to_integral() else float(number_of_cuts)
            existing_invoice.peak_quantity = float(peak_quantity)
            existing_invoice.peak_width = float(peak_width)
            existing_invoice.lamination_cost = float(lamination_cost)
            # preserve your model naming (you used Fee earlier)
            setattr(existing_invoice, 'Fee', float(fee))
            existing_invoice.row_number = row_number
            existing_invoice.total_price = float(total_price)
            existing_invoice.notes = notes

            if existing_draft:
                db.session.delete(existing_draft)

            db.session.commit()

            # mark order invoiced
            try:
                order.invoiced = True
                db.session.commit()
            except Exception:
                db.session.rollback()

            return {
                "success": True,
                "message": "Invoice updated successfully from factory processing",
                "invoice_number": existing_invoice.invoice_number,
                "total_price": float(total_price),
                "is_update": True
            }

        # Create new invoice
        last_invoice = Payment.query.order_by(Payment.id.desc()).first()
        new_invoice_number_id = (last_invoice.id if last_invoice else 0) + 1
        invoice_number = f"INV-{datetime.utcnow().year}-{new_invoice_number_id:03d}"

        new_invoice = Payment(
            order_id=order_id,
            invoice_number=invoice_number,
            unit_price=float(unit_price),
            quantity=int(quantity) if quantity == quantity.to_integral() else float(quantity),
            cutting_cost=float(cutting_cost),
            number_of_cuts=int(number_of_cuts) if number_of_cuts == number_of_cuts.to_integral() else float(number_of_cuts),
            peak_quantity=float(peak_quantity),
            peak_width=float(peak_width),
            lamination_cost=float(lamination_cost),
            Fee=float(fee),
            row_number=row_number,
            total_price=float(total_price),
            status='Generated',
            notes=notes,
            created_by=user_id
        )

        db.session.add(new_invoice)
        if existing_draft:
            db.session.delete(existing_draft)

        try:
            order.invoiced = True
        except Exception:
            pass

        db.session.commit()

        return {
            "success": True,
            "message": "Invoice created successfully from factory processing",
            "invoice_number": invoice_number,
            "total_price": float(total_price),
            "is_update": False
        }

    except Exception as e:
        db.session.rollback()
        return {"success": False, "message": f"Error saving factory invoice: {e}"}

def save_invoice_draft(order_id: int, invoice_data: Dict[str, Any], user_id: int) -> Dict[str, Any]:
    """
    Save invoice data as a draft for later completion when order is finished.
    Now supports 'lamination_cost' (added to InvoiceDraft table).
    """
    try:
        # Check if draft already exists for this order
        existing_draft = InvoiceDraft.query.filter_by(order_id=order_id).first()
        order = Order.query.get(order_id)
        if not order:
            return False, {"error": "Order not found"}

        # Extract invoice data (now includes lamination_cost)
        quantity = invoice_data.get('quantity', 0)
        cutting_cost = invoice_data.get('cutting_cost', 0.0)
        number_of_cuts = invoice_data.get('number_of_cuts', 0)
        peak_quantity = invoice_data.get('peak_quantity', 0.0)
        peak_width = invoice_data.get('peak_width', 0.0)
        fee = invoice_data.get('Fee', invoice_data.get('fee', 0.0))
        lamination_cost = invoice_data.get('lamination_cost')
        row_number = invoice_data.get('row_number')
        notes = invoice_data.get('notes', '')

        # Ensure all numeric inputs are cast correctly
        try:
            quantity = int(quantity) if quantity else 0
            cutting_cost = float(cutting_cost) if cutting_cost else 0.0
            number_of_cuts = int(number_of_cuts) if number_of_cuts else 0
            peak_quantity = float(peak_quantity) if peak_quantity else 0.0
            peak_width = float(peak_width) if peak_width else 0.0
            fee = float(fee) if fee else 0.0
            lamination_cost = float(lamination_cost) if lamination_cost else 0.0
            row_number = int(row_number) if row_number else None
        except (ValueError, TypeError) as e:
            return {"success": False, "message": f"Invalid numeric value: {e}"}

        # Optionally update some order fields (as you did before)
        if quantity is not None:
            try:
                order.produced_quantity = float(quantity)
            except Exception:
                return {"success": False, "message": "Invalid quantity value"}

        if fee is not None:
            try:
                if getattr(order, 'customer', None):
                    order.customer.fee = float(fee)
            except Exception:
                return {"success": False, "message": "Invalid fee value"}

        if peak_width is not None:
            try:
                order.width = float(peak_width)
            except Exception:
                return {"success": False, "message": "Invalid width value"}

        if existing_draft:
            # Update existing draft (include lamination_cost)
            existing_draft.quantity = quantity
            existing_draft.cutting_cost = cutting_cost
            existing_draft.number_of_cuts = number_of_cuts
            existing_draft.peak_quantity = peak_quantity
            existing_draft.peak_width = peak_width
            existing_draft.Fee = fee
            existing_draft.lamination_cost = lamination_cost
            existing_draft.row_number = row_number
            existing_draft.notes = notes
            existing_draft.updated_at = datetime.now(timezone.utc)

            db.session.commit()

            return {
                "success": True,
                "message": "Invoice draft updated successfully",
                "is_update": True
            }
        else:
            # Create new draft (include lamination_cost)
            new_draft = InvoiceDraft(
                order_id=order_id,
                quantity=quantity,
                cutting_cost=cutting_cost,
                lamination_cost=lamination_cost,
                number_of_cuts=number_of_cuts,
                peak_quantity=peak_quantity,
                peak_width=peak_width,
                Fee=fee,
                row_number=row_number,
                notes=notes,
                created_by=user_id
            )

            db.session.add(new_draft)
            db.session.commit()

            return {
                "success": True,
                "message": "Invoice draft created successfully",
                "is_update": False
            }

    except Exception as e:
        db.session.rollback()
        return {"success": False, "message": f"Error saving invoice draft: {e}"}

def get_cutting_price_for_sketch(sketch_name: str) -> float:
    """
    Returns the cutting price based on the prefix of sketch_name.
    باریک = 50, اتیکت = 75, کتی = 90
    """
    if not sketch_name:
        return 0.0
    sketch_name = sketch_name.strip()
    if sketch_name.startswith("باریک"):
        return 50.0
    elif sketch_name.startswith("اتیکت"):
        return 75.0
    elif sketch_name.startswith("کتی"):
        return 90.0
    return 0.0


def get_number_of_cuts_for_order(order_id: int) -> int:
    """
    Returns the number_of_cuts for the given order,
    based on the member_count of ProductionStepLog where step_name is BRESH.
    If multiple logs exist, returns the sum.
    """
    print(f"started get number of cuts function ")
    logs = ProductionStepLog.query.filter_by(
        order_id=order_id,
        step_name="bresh"
    ).all()
    return sum(log.member_count for log in logs if log.member_count is not None)
PRESS_UNIT_COST = Decimal('350')  # one-time press fee if condition met

def _compute_press_cost_one_time(order_id: int) -> Decimal:
    """
    Return PRESS_UNIT_COST if at least one ProductionStepLog for this order
    indicates a press step with member_count > 0. Robust to:
     - step_name stored as enum or string (case-insensitive, substring match)
     - member_count stored as int, numeric-string, or float-string
    """
    try:
        logs = ProductionStepLog.query.filter_by(order_id=order_id).all()
        for log in logs:
            step_raw = getattr(log, "step_name", "") or ""
            # normalize step_name whether enum or string
            try:
                # if it's an enum value (SQLAlchemy Enum), use .name or str()
                step_str = (step_raw.name if hasattr(step_raw, "name") else str(step_raw)).lower()
            except Exception:
                step_str = str(step_raw).lower()

            # accept 'press' anywhere in name (helps with different conventions)
            if "press" not in step_str:
                continue

            # parse member_count defensively
            member_raw = getattr(log, "member_count", 0) or 0
            try:
                member_count = int(member_raw)
            except Exception:
                try:
                    member_count = int(float(member_raw))
                except Exception:
                    member_count = 0

            if member_count > 0:
                return PRESS_UNIT_COST

        return Decimal("0")
    except Exception:
        # be conservative on DB errors
        return Decimal("0")
def get_number_of_cuts_for_order(order_id: int) -> int:
    """
    Returns the number_of_cuts for the given order,
    based on the member_count of ProductionStepLog where step_name is BRESH.
    If multiple logs exist, returns the sum.
    """
    logs = ProductionStepLog.query.filter_by(
        order_id=order_id,
        step_name="bresh"
    ).all()
    return sum(log.member_count for log in logs if log.member_count is not None)
