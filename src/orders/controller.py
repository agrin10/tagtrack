from src import db
from src.orders.models import Order
from flask_login import current_user
from datetime import datetime, date
from typing import Tuple, Dict, Any, List
import traceback
from sqlalchemy import func
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

def _get_next_form_number_for_year() -> int:
    current_year = datetime.now().year
    start_of_year = datetime(current_year, 1, 1)
    end_of_year = datetime(current_year, 12, 31, 23, 59, 59)

    # Query for the maximum form_number within the current year
    max_form_number = db.session.query(func.max(Order.form_number)).filter(
        Order.created_at.between(start_of_year, end_of_year)
    ).scalar()

    return (max_form_number or 0) + 1

def add_order(form_data: Dict[str, Any]) -> Tuple[bool, Dict[str, Any]]:
    try:
        # Debug print the incoming data
        print("Raw form data received:", form_data)
        
        # Convert all string values to str type and strip whitespace
        form_data = {k: str(v).strip() if isinstance(v, (str, int, float)) else v 
                    for k, v in form_data.items() if v is not None}
        
        print("Processed form data:", form_data)

        # Customer name is required
        if not form_data.get('customer_name'):
            return False, {"error": "Customer name is required"}

        # Auto-generate form_number with yearly reset
        form_number = _get_next_form_number_for_year()

        # Parse dates if provided
        delivery_date = None
        if form_data.get('delivery_date'):
            try:
                delivery_date = datetime.strptime(form_data['delivery_date'], '%Y-%m-%d').date()
            except ValueError:
                return False, {"error": "Invalid delivery date format. Use YYYY-MM-DD"}

        # Parse exit dates if provided
        exit_from_office_date = None
        if form_data.get('exit_from_office_date'):
            try:
                exit_from_office_date = datetime.strptime(form_data['exit_from_office_date'], '%Y-%m-%d').date()
            except ValueError:
                return False, {"error": "Invalid exit from office date format. Use YYYY-MM-DD"}

        exit_from_factory_date = None
        if form_data.get('exit_from_factory_date'):
            try:
                exit_from_factory_date = datetime.strptime(form_data['exit_from_factory_date'], '%Y-%m-%d').date()
            except ValueError:
                return False, {"error": "Invalid exit from factory date format. Use YYYY-MM-DD"}

        # Convert numeric fields
        try:
            width = float(form_data['width']) if form_data.get('width') else None
            height = float(form_data['height']) if form_data.get('height') else None
            quantity = int(form_data['quantity']) if form_data.get('quantity') else None
            total_length_meters = float(form_data['total_length_meters']) if form_data.get('total_length_meters') else None
        except ValueError as e:
            return False, {"error": f"Invalid numeric value: {str(e)}"}

        # Create new order
        new_order = Order(
            form_number=form_number, # Use the auto-generated form number
            customer_name=form_data['customer_name'],
            fabric_density=form_data.get('fabric_density'),
            fabric_cut=form_data.get('fabric_cut'),
            width=width,
            height=height,
            quantity=quantity,
            total_length_meters=total_length_meters,
            delivery_date=delivery_date,
            exit_from_office_date=exit_from_office_date,
            exit_from_factory_date=exit_from_factory_date,
            sketch_name=form_data.get('sketch_name'),
            file_name=form_data.get('file_name'),
            design_specification=form_data.get('design_specification'),
            office_notes=form_data.get('office_notes'),
            factory_notes=form_data.get('factory_notes'),
            customer_note_to_office=form_data.get('customer_note_to_office'),
            fusing_type=form_data.get('fusing_type'),
            lamination_type=form_data.get('lamination_type'),
            cut_type=form_data.get('cut_type'),
            label_type=form_data.get('label_type'),
            created_by=current_user.id,
            status=form_data.get('status', 'Pending'),  # Default status to 'Pending'
        )

        # Add to database
        db.session.add(new_order)
        db.session.commit()

        # Refresh the order to get the database-generated values
        db.session.refresh(new_order)

        return True, {
            "message": "Order created successfully",
            "order": new_order.to_dict()
        }

    except Exception as e:
        db.session.rollback()
        # Get the full traceback
        error_traceback = traceback.format_exc()
        print("Error creating order:")
        print(error_traceback)
        # Return the actual error message for debugging
        return False, {"error": f"Failed to create order: {str(e)}\nTraceback: {error_traceback}"}

def get_orders(page: int = 1, per_page: int = 10, search: str = None, status: str = None) -> Tuple[bool, Dict[str, Any]]:
    """
    Get all orders with their details, with pagination and optional search/status filters.
    Returns a tuple of (success, response) where response contains either the orders list or an error message.
    """
    try:
        query = Order.query

        # Apply search filter
        if search:
            query = query.filter(
                db.or_(
                    Order.customer_name.ilike(f'%{search}%'),
                    db.cast(Order.form_number, db.String).ilike(f'%{search}%')
                )
            )
        
        # Apply status filter
        if status and status.lower() != 'all':
            query = query.filter(Order.status == status)

        # Order by created_at descending (newest first)
        query = query.order_by(Order.created_at.desc())
        
        # Paginate the results
        pagination = query.paginate(page=page, per_page=per_page, error_out=False)
        
        # Convert orders to list of dictionaries
        orders_list = [order.to_dict() for order in pagination.items]
        
        return True, {
            "message": "Orders retrieved successfully",
            "orders": orders_list,
            "total": pagination.total,
            "pagination": pagination
        }
        
    except Exception as e:
        print(f"Error retrieving orders: {str(e)}")
        return False, {"error": "Failed to retrieve orders"}
    
def get_order_by_id(order_id: int) -> Tuple[bool, Dict[str, Any]]:
    """
    Get a specific order by its ID.
    """
    try:
        order = Order.query.get(order_id)
        if not order:
            return False, {"error": "Order not found"}
        
        return True, {
            "message": "Order retrieved successfully",
            "order": order.to_dict()
        }
        
    except Exception as e:
        print(f"Error retrieving order {order_id}: {str(e)}")
        return False, {"error": f"Failed to retrieve order: {str(e)}"}
    
def delete_order_by_id(order_id: int) -> Tuple[bool , Dict[str, Any]]:
    """
    Delete an order by its ID.
    """
    try:
        order = Order.query.get(order_id)
        if not order:
            return False, {"error": "Order not found"}
        
        db.session.delete(order)
        db.session.commit()
        
        return True, {
            "message": "Order deleted successfully"
        }
        
    except Exception as e:
        db.session.rollback()
        print(f"Error deleting order {order_id}: {str(e)}")
        return False, {"error": f"Failed to delete order: {str(e)}"}
    
def update_order_id(order_id: int, form_data: Dict[str, Any]) -> Tuple[bool, Dict[str, Any]]:
    """
    Update an existing order with the provided form data.
    """
    try:
        print(f"Starting update for order {order_id}")
        print("Received form data:", form_data)
        
        order = Order.query.get(order_id)
        if not order:
            print(f"Order {order_id} not found")
            return False, {"error": "Order not found"}
        
        print("Found order:", order.to_dict())
        
        # Update fields from form_data
        for key, value in form_data.items():
            print(f"Processing field {key} with value {value} (type: {type(value)})")
            if hasattr(order, key):
                if key == 'created_at' and value:
                    print(f"Processing created_at value: {value} (type: {type(value)})")
                    try:
                        # Convert ISO string to datetime
                        value = datetime.fromisoformat(value.replace('Z', '+00:00'))
                        print(f"Converted created_at to datetime: {value}")
                    except ValueError as e:
                        print(f"Error converting created_at: {str(e)}")
                        return False, {"error": f"Invalid created_at datetime format: {str(e)}"}
                elif key == 'delivery_date' and value:
                    try:
                        # Convert ISO string to date
                        value = datetime.fromisoformat(value).date()
                        print(f"Converted delivery_date to date: {value}")
                    except ValueError as e:
                        print(f"Error converting delivery_date: {str(e)}")
                        return False, {"error": f"Invalid delivery_date format: {str(e)}"}
                elif key in ['exit_from_office_date', 'exit_from_factory_date'] and value:
                    try:
                        # Convert ISO string to date
                        value = datetime.fromisoformat(value).date()
                        print(f"Converted {key} to date: {value}")
                    except ValueError as e:
                        print(f"Error converting {key}: {str(e)}")
                        return False, {"error": f"Invalid {key} format: {str(e)}"}
                print(f"Setting {key} to {value}")
                setattr(order, key, value)
            else:
                print(f"Field {key} not found in Order model")
        
        # Update timestamps
        order.updated_at = datetime.utcnow()
        print("Updated timestamps")
        
        print("Committing changes to database")
        db.session.commit()
        print("Changes committed successfully")
        
        updated_order = order.to_dict()
        print("Updated order data:", updated_order)
        
        return True, {
            "message": "Order updated successfully",
            "order": updated_order
        }
        
    except Exception as e:
        db.session.rollback()
        print(f"Error updating order {order_id}: {str(e)}")
        print("Full error details:", traceback.format_exc())
        return False, {"error": f"Failed to update order: {str(e)}"}

def duplicate_order(order_id):
    """
    Duplicate an existing order with a new ID.
    """
    try:
        # Get the original order
        success, response = get_order_by_id(order_id)
        if not success:
            return False, {"error": "Order not found"}

        original_order = response.get('order')
        if not original_order:
            return False, {"error": "Order not found"}

        # Auto-generate form_number with yearly reset
        form_number = _get_next_form_number_for_year()

        # Create a copy of the order data with a new form number
        new_order_data = {
            'form_number': form_number,  # Use the auto-generated form number
            'customer_name': original_order.get('customer_name'),
            'fabric_density': original_order.get('fabric_density'),
            'fabric_cut': original_order.get('fabric_cut'),
            'width': original_order.get('width'),
            'height': original_order.get('height'),
            'quantity': original_order.get('quantity'),
            'total_length_meters': original_order.get('total_length_meters'),
            'fusing_type': original_order.get('fusing_type'),
            'lamination_type': original_order.get('lamination_type'),
            'cut_type': original_order.get('cut_type'),
            'label_type': original_order.get('label_type'),
            'delivery_date': original_order.get('delivery_date'),
            'exit_from_office_date': original_order.get('exit_from_office_date'),
            'exit_from_factory_date': original_order.get('exit_from_factory_date'),
            'sketch_name': original_order.get('sketch_name'),
            'file_name': original_order.get('file_name'),
            'status': 'Pending',  # Set status to Pending for the new order
            'design_specification': original_order.get('design_specification'),
            'office_notes': original_order.get('office_notes'),
            'factory_notes': original_order.get('factory_notes'),
            'customer_note_to_office': original_order.get('customer_note_to_office')
        }

        # Add the new order
        success, response = add_order(new_order_data)
        if success:
            return True, {"order": response.get('order')}
        return False, response

    except Exception as e:
        print(f"Error in duplicate_order: {str(e)}")
        return False, {"error": "An error occurred while duplicating the order"}

def generate_excel_report(search: str = None, status: str = None) -> Tuple[bool, Dict[str, Any]]:
    """
    Generate an Excel report of orders, with optional search/status filters.
    Returns a tuple of (success, response) where response contains either the Excel file buffer or an error message.
    """
    try:
        query = Order.query

        # Apply search filter
        if search:
            query = query.filter(
                db.or_(
                    Order.customer_name.ilike(f'%{search}%'),
                    db.cast(Order.form_number, db.String).ilike(f'%{search}%')
                )
            )
        
        # Apply status filter
        if status and status.lower() != 'all':
            query = query.filter(Order.status == status)

        # Order by created_at descending (newest first)
        orders = query.order_by(Order.created_at.desc()).all()

        # Create a new workbook and select the active worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Orders Report"

        # Define headers and their corresponding order attributes
        headers = [
            "Form Number", "Customer Name", "Sketch Name", "File Name",
            "Fabric Density", "Fabric Cut", "Width", "Height", "Quantity", 
            "Total Length (m)", "Delivery Date", "Exit from Office Date",
            "Exit from Factory Date", "Print Type", "Lamination Type", 
            "Cut Type", "Label Type", "Design Specification", "Office Notes",
            "Factory Notes", "Customer Note to Office", "Status", 
            "Created At", "Updated At", "Created By"
        ]
        order_attributes = [
            "form_number", "customer_name", "sketch_name", "file_name",
            "fabric_density", "fabric_cut", "width", "height", "quantity",
            "total_length_meters", "delivery_date", "exit_from_office_date",
            "exit_from_factory_date", "fusing_type", "lamination_type",
            "cut_type", "label_type", "design_specification", "office_notes",
            "factory_notes", "customer_note_to_office", "status",
            "created_at", "updated_at", "created_by_username"
        ]

        # Write headers
        ws.append(headers)

        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_border = Border(left=Side(style='thin'), 
                               right=Side(style='thin'), 
                               top=Side(style='thin'), 
                               bottom=Side(style='thin'))

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = header_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        # Write order data
        for order in orders:
            row_data = []
            order_dict = order.to_dict() # Get the dictionary representation of the order
            for attr in order_attributes:
                value = order_dict.get(attr)
                
                # Format dates to YYYY-MM-DD
                if isinstance(value, (datetime, date)):
                    value = value.strftime('%Y-%m-%d')
                elif isinstance(value, str) and 'T' in value and (
                    attr == 'created_at' or attr == 'updated_at'):
                    # Handle ISO format strings from to_dict()
                    try:
                        value = datetime.fromisoformat(value).strftime('%Y-%m-%d')
                    except ValueError:
                        pass # Keep original string if parsing fails
                
                row_data.append(value)
            ws.append(row_data)

        # Auto-size columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2 # Add a small buffer
            ws.column_dimensions[column].width = adjusted_width

        # Save workbook to a BytesIO object
        excel_file_buffer = io.BytesIO()
        wb.save(excel_file_buffer)
        excel_file_buffer.seek(0) # Rewind the buffer to the beginning
        
        current_date = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_name = f"orders_report_{current_date}.xlsx"

        return True, {
            "message": "Excel report generated successfully",
            "excel_file_buffer": excel_file_buffer,
            "file_name": file_name
        }

    except Exception as e:
        db.session.rollback()
        error_traceback = traceback.format_exc()
        print("Error generating Excel report:")
        print(error_traceback)
        return False, {"error": f"Failed to generate Excel report: {str(e)}\nTraceback: {error_traceback}"}